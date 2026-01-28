from robot.api import ExecutionResult
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image
from datetime import datetime
import os

# =========================
# Load Robot output.xml
# =========================
result = ExecutionResult("results/output.xml")

rows = []

def collect_tests(suite):
    for test in suite.tests:
        rows.append({
            "Test Name": test.name,
            "Status": test.status,
            "Start Time": test.starttime,
            "End Time": test.endtime,
            "Message": test.message
        })
    for sub in suite.suites:
        collect_tests(sub)

collect_tests(result.suite)

df = pd.DataFrame(rows)

# =========================
# Create Excel file
# =========================
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_path = f"results/results_{timestamp}.xlsx"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Results", index=False)

wb = load_workbook(excel_path)
ws = wb["Results"]

# =========================
# Styles
# =========================
header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
header_font = Font(bold=True)

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# =========================
# Header style (Results)
# =========================
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

# =========================
# Status coloring
# =========================
status_col = 2  # Status column (B)

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=status_col)
    if cell.value == "PASS":
        cell.fill = green
    elif cell.value == "FAIL":
        cell.fill = red

# =========================
# Add Screenshot column
# =========================
screenshot_col = ws.max_column + 1
ws.cell(row=1, column=screenshot_col).value = "Screenshot"
ws.cell(row=1, column=screenshot_col).fill = header_fill
ws.cell(row=1, column=screenshot_col).font = header_font

# =========================
# Summary sheet
# =========================
summary = wb.create_sheet("Summary")

summary["A1"] = "Metric"
summary["B1"] = "Count"
summary["A2"] = "Total Tests"
summary["B2"] = len(df)
summary["A3"] = "Passed"
summary["B3"] = len(df[df["Status"] == "PASS"])
summary["A4"] = "Failed"
summary["B4"] = len(df[df["Status"] == "FAIL"])

for col in range(1, 3):
    summary.cell(row=1, column=col).fill = header_fill
    summary.cell(row=1, column=col).font = header_font

summary["B3"].fill = green
summary["B4"].fill = red

# =========================
# Screenshot mapping
# =========================
if "Screenshots" in wb.sheetnames:
    del wb["Screenshots"]

shot_sheet = wb.create_sheet("Screenshots")
shot_sheet["A1"] = "Test Case"
shot_sheet["B1"] = "Screenshot"

shot_sheet["A1"].fill = header_fill
shot_sheet["B1"].fill = header_fill
shot_sheet["A1"].font = header_font
shot_sheet["B1"].font = header_font

# Collect all screenshots
screenshot_folder = "results"
screenshots = {}

for file in os.listdir(screenshot_folder):
    if file.lower().endswith(".png"):
        screenshots[file.lower()] = os.path.join(screenshot_folder, file)

img_row = 2


# =========================
# Link FAIL test → Screenshot
# =========================
for row in range(2, ws.max_row + 1):
    test_name = ws.cell(row=row, column=1).value
    status = ws.cell(row=row, column=2).value

    if status == "FAIL":
        matched_img = None

        for file, path in screenshots.items():
            if test_name.lower().replace(" ", "-") in file:
                matched_img = path
                break

        if matched_img:
            # Hyperlink in Results
            link_cell = ws.cell(row=row, column=screenshot_col)
            link_cell.value = "View Screenshot"
            link_cell.hyperlink = f"#Screenshots!A{img_row}"
            link_cell.style = "Hyperlink"

            # Add image to Screenshots sheet
            shot_sheet.cell(row=img_row, column=1).value = test_name

            img = Image(matched_img)
            img.width = 400
            img.height = 250
            shot_sheet.add_image(img, f"B{img_row}")

            img_row += 15
# =========================
# Link ALL tests → Screenshot (PASS & FAIL)
# =========================
screenshot_list = list(screenshots.values())  # เรียงตามลำดับไฟล์
shot_index = 0

for row in range(2, ws.max_row + 1):
    if shot_index >= len(screenshot_list):
        break

    test_name = ws.cell(row=row, column=1).value
    status = ws.cell(row=row, column=2).value
    matched_img = screenshot_list[shot_index]

    # Link in Results
    link_cell = ws.cell(row=row, column=screenshot_col)
    link_cell.value = "View Screenshot"
    link_cell.hyperlink = f"#Screenshots!A{img_row}"
    link_cell.style = "Hyperlink"

    # Add image to Screenshots sheet
    shot_sheet.cell(row=img_row, column=1).value = f"{test_name} ({status})"

    img = Image(matched_img)
    img.width = 400
    img.height = 250
    shot_sheet.add_image(img, f"B{img_row}")

    img_row += 15
    shot_index += 1

# =========================
# Save Excel
# =========================
wb.save(excel_path)

print("✅ 📷 Excel report created successfully:", excel_path)
